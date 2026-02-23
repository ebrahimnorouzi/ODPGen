#!/usr/bin/env python3
"""Aggregate blinded human/LLM evaluation scores and generate summary artifacts."""

from __future__ import annotations

import argparse
import csv
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

DIMENSIONS = ["correctness_1_to_5", "completeness_1_to_5", "clarity_reusability_1_to_5"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--ratings", nargs="+", required=True, help="Completed rating packet files")
    parser.add_argument("--mapping", default="results/human_eval_mapping.csv", help="Anonymized ID mapping CSV")
    parser.add_argument("--results-dir", default="results", help="Output directory")
    parser.add_argument(
        "--rater-column",
        default="rater_id",
        help="Column name containing evaluator identifier (e.g., rater_id or llm_model)",
    )
    parser.add_argument(
        "--evaluator-type",
        default="mixed",
        help="Label for this run's evaluator type (e.g., human, llm, mixed)",
    )
    return parser.parse_args()


def _normalize_value(raw: object) -> Optional[float]:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    return value if 1 <= value <= 5 else None


def read_table(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8", newline="") as handle:
            return list(csv.DictReader(handle))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(f"openpyxl is required to parse {path}. Use CSV or install openpyxl.") from exc
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(v).strip() if v is not None else "" for v in rows[0]]
        out: List[Dict[str, str]] = []
        for line in rows[1:]:
            row: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = "" if idx >= len(line) or line[idx] is None else str(line[idx])
                row[header] = value
            out.append(row)
        return out
    raise ValueError(f"Unsupported rating file format: {path}")


def load_mapping(mapping_path: Path) -> Dict[str, Dict[str, str]]:
    with mapping_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    mapping: Dict[str, Dict[str, str]] = {}
    for row in rows:
        key = row.get("anonymized_output_id", "").strip()
        if key:
            mapping[key] = row
    return mapping


def rater_from_path(path: Path) -> str:
    return path.stem.rsplit("__", 1)[-1] if "__" in path.stem else path.stem


def safe_mean(values: Iterable[float]) -> Optional[float]:
    vals = list(values)
    return statistics.mean(vals) if vals else None


def safe_median(values: Iterable[float]) -> Optional[float]:
    vals = list(values)
    return statistics.median(vals) if vals else None


def pct(numerator: int, denominator: int) -> Optional[float]:
    return (numerator / denominator) * 100.0 if denominator else None


def _round_or_blank(value: Optional[float]) -> object:
    return round(value, 4) if isinstance(value, float) else ""


def compute_agreement(rows: Sequence[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    agreement_stats: Dict[str, Dict[str, object]] = {}
    for dim in DIMENSIONS:
        comparable = 0
        exact = 0
        spread_values: List[float] = []

        by_output: Dict[str, List[float]] = defaultdict(list)
        for row in rows:
            val = row.get(dim)
            if isinstance(val, float):
                by_output[str(row["anonymized_output_id"])].append(val)

        for vals in by_output.values():
            if len(vals) < 2:
                continue
            comparable += 1
            if len(set(vals)) == 1:
                exact += 1
            spread_values.append(statistics.pstdev(vals))

        agreement_stats[dim] = {
            "comparable_outputs": comparable,
            "exact_agreement_pct": _round_or_blank(pct(exact, comparable)),
            "avg_rating_stddev": _round_or_blank(statistics.mean(spread_values) if spread_values else None),
            "krippendorff_alpha_placeholder": "TODO",
        }
    return agreement_stats


def main() -> int:
    args = parse_args()
    results_dir = Path(args.results_dir)
    results_dir.mkdir(parents=True, exist_ok=True)

    mapping = load_mapping(Path(args.mapping))
    if not mapping:
        raise SystemExit("Mapping file is empty or missing anonymized_output_id rows.")

    score_rows: List[Dict[str, object]] = []
    for rating_file in [Path(p) for p in args.ratings]:
        table = read_table(rating_file)
        fallback_rater = rater_from_path(rating_file)
        for row in table:
            anon_id = str(row.get("anonymized_output_id", "")).strip()
            if not anon_id or anon_id not in mapping:
                continue
            evaluator_id = str(row.get(args.rater_column, "")).strip() or fallback_rater
            mapping_row = mapping[anon_id]
            score_rows.append(
                {
                    "anonymized_output_id": anon_id,
                    "scenario_id": mapping_row.get("scenario_id", ""),
                    "model": mapping_row.get("model", ""),
                    "config": mapping_row.get("config", ""),
                    "evaluator_type": args.evaluator_type,
                    "evaluator_id": evaluator_id,
                    **{dim: _normalize_value(row.get(dim)) for dim in DIMENSIONS},
                }
            )

    if not score_rows:
        raise SystemExit("No valid scored rows found in provided rating files.")

    grouped: Dict[Tuple[str, str], List[Dict[str, object]]] = defaultdict(list)
    for row in score_rows:
        grouped[(str(row["model"]), str(row["config"]))].append(row)

    summary_rows: List[Dict[str, object]] = []
    for (model, config), rows in sorted(grouped.items()):
        summary: Dict[str, object] = {
            "evaluator_type": args.evaluator_type,
            "model": model,
            "config": config,
            "n_scored_rows": len(rows),
            "n_unique_outputs": len({str(r["anonymized_output_id"]) for r in rows}),
            "n_unique_evaluators": len({str(r["evaluator_id"]) for r in rows}),
        }
        for dim in DIMENSIONS:
            values = [v for v in (r.get(dim) for r in rows) if isinstance(v, float)]
            summary[f"{dim}_mean"] = _round_or_blank(safe_mean(values))
            summary[f"{dim}_median"] = _round_or_blank(safe_median(values))
        summary_rows.append(summary)

    agreement_stats = compute_agreement(score_rows)

    summary_path = results_dir / "human_eval_summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(summary_rows[0].keys()))
        writer.writeheader()
        writer.writerows(summary_rows)

    agreement_path = results_dir / "human_eval_agreement.csv"
    with agreement_path.open("w", newline="", encoding="utf-8") as handle:
        fields = [
            "evaluator_type",
            "dimension",
            "comparable_outputs",
            "exact_agreement_pct",
            "avg_rating_stddev",
            "krippendorff_alpha_placeholder",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for dim in DIMENSIONS:
            writer.writerow({"evaluator_type": args.evaluator_type, "dimension": dim, **agreement_stats[dim]})

    report_path = results_dir / "human_eval_summary.md"
    lines = [
        "# Human/LLM Evaluation Summary",
        "",
        f"Evaluator type: **{args.evaluator_type}**",
        f"Scored rows: **{len(score_rows)}**",
        f"Unique outputs: **{len({str(r['anonymized_output_id']) for r in score_rows})}**",
        f"Unique evaluators: **{len({str(r['evaluator_id']) for r in score_rows})}**",
        "",
        "## Mean / Median by Model + Config",
        "",
        "| Model | Config | Correctness (mean/median) | Completeness (mean/median) | Clarity-Reusability (mean/median) |",
        "|---|---|---:|---:|---:|",
    ]
    for row in summary_rows:
        lines.append(
            "| {model} | {config} | {cm}/{cd} | {pm}/{pd} | {lm}/{ld} |".format(
                model=row["model"],
                config=row["config"],
                cm=row.get("correctness_1_to_5_mean", ""),
                cd=row.get("correctness_1_to_5_median", ""),
                pm=row.get("completeness_1_to_5_mean", ""),
                pd=row.get("completeness_1_to_5_median", ""),
                lm=row.get("clarity_reusability_1_to_5_mean", ""),
                ld=row.get("clarity_reusability_1_to_5_median", ""),
            )
        )

    lines.extend([
        "",
        "## Inter-rater Agreement (available stats)",
        "",
        "| Dimension | Comparable Outputs (>=2 evaluators) | Exact Agreement % | Avg Rating StdDev | Krippendorff Alpha |",
        "|---|---:|---:|---:|---|",
    ])
    for dim in DIMENSIONS:
        stat = agreement_stats[dim]
        lines.append(
            f"| {dim} | {stat['comparable_outputs']} | {stat['exact_agreement_pct']} | {stat['avg_rating_stddev']} | {stat['krippendorff_alpha_placeholder']} |"
        )
    lines.append("")
    report_path.write_text("\n".join(lines), encoding="utf-8")

    print(f"Wrote summary CSV: {summary_path}")
    print(f"Wrote agreement CSV: {agreement_path}")
    print(f"Wrote summary report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
